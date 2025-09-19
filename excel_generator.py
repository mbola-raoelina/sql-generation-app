"""
Excel Generator Module
Handles Excel report generation with SOAP and Selenium fallback.
"""

import streamlit as st
import time
import os
import logging
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import requests
from zeep import Client
from zeep.transports import Transport
from requests import Session

# Configure logging for Excel generation
excel_logger = logging.getLogger('excel_generator')
excel_logger.setLevel(logging.INFO)

# Create a dedicated Excel generation log file
excel_log_file = f'excel_debug_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'

# Clear any existing handlers for this specific logger
for handler in excel_logger.handlers[:]:
    excel_logger.removeHandler(handler)

# Create file handler for Excel generation logs
excel_file_handler = logging.FileHandler(excel_log_file, mode='w')
excel_file_handler.setLevel(logging.INFO)

# Create console handler
excel_console_handler = logging.StreamHandler()
excel_console_handler.setLevel(logging.INFO)

# Create formatter
excel_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
excel_file_handler.setFormatter(excel_formatter)
excel_console_handler.setFormatter(excel_formatter)

# Add handlers to the logger
excel_logger.addHandler(excel_file_handler)
excel_logger.addHandler(excel_console_handler)

excel_logger.info("=== EXCEL GENERATOR LOGGING INITIALIZED ===")
excel_logger.info(f"Excel log file: {excel_log_file}")

# Use the dedicated Excel logger
logger = excel_logger

class ExcelGenerator:
    """Handles Excel report generation with multiple approaches."""
    
    def __init__(self):
        # Load environment variables
        import os
        from dotenv import load_dotenv
        load_dotenv()
        
        self.bip_endpoint = os.environ.get('BIP_ENDPOINT')
        self.bip_username = os.environ.get('BIP_USERNAME')
        self.bip_password = os.environ.get('BIP_PASSWORD')
        self.bip_report_path = os.environ.get('BIP_REPORT_PATH')
        self.bip_wsdl_url = os.environ.get('BIP_WSDL_URL')
        
        logger.info("ExcelGenerator initialized")
    
    def generate_excel_via_soap(self, sql_query: str) -> tuple[bytes, int]:
        """Generate Excel report via SOAP API."""
        try:
            logger.info("Attempting SOAP-based Excel generation...")
            
            if not all([self.bip_endpoint, self.bip_username, self.bip_password, self.bip_wsdl_url]):
                logger.warning("Missing BI Publisher configuration")
                return None, 0
            
            # Remove semicolon at the end as BI Publisher doesn't accept it
            safe_sql = sql_query.rstrip(';').strip()
            if safe_sql != sql_query:
                logger.info("Removed semicolon from SQL query for BI Publisher compatibility")
            
            # If legacy modules are unavailable, build the same SOAP envelope and POST directly
            try:
                import base64
                import gzip
                from io import BytesIO
                import xml.etree.ElementTree as ET

                # GZIP + base64 encode the SQL (as used by appli)
                buf = BytesIO()
                with gzip.GzipFile(fileobj=buf, mode='wb', mtime=0) as gz:
                    gz.write(safe_sql.encode('utf-8'))
                b64_sql = base64.b64encode(buf.getvalue()).decode('utf-8')

                # Build SOAP envelope mirroring appli.py's create_soap_payload (v2 ReportService)
                # Decode URL-encoded characters in report path for SOAP service
                import urllib.parse
                report_path = urllib.parse.unquote(self.bip_report_path)
                
                # Use the exact working SOAP envelope from our debug script
                envelope = f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:v2="http://xmlns.oracle.com/oxp/service/v2">
  <soapenv:Header/>
  <soapenv:Body>
    <v2:runReport>
      <v2:reportRequest>
        <v2:reportAbsolutePath>{report_path}</v2:reportAbsolutePath>
        <v2:attributeFormat>csv</v2:attributeFormat>
        <v2:sizeOfDataChunkDownload>-1</v2:sizeOfDataChunkDownload>
        <v2:parameterNameValues>
          <v2:listOfParamNameValues>
            <v2:item>
              <v2:name>P_B64_CONTENT</v2:name>
              <v2:values><v2:item>{b64_sql}</v2:item></v2:values>
            </v2:item>
          </v2:listOfParamNameValues>
        </v2:parameterNameValues>
      </v2:reportRequest>
      <v2:userID>{self.bip_username}</v2:userID>
      <v2:password>{self.bip_password}</v2:password>
    </v2:runReport>
  </soapenv:Body>
</soapenv:Envelope>"""

                # POST to the same endpoint path as appli
                service_url = f"{self.bip_endpoint}:443/xmlpserver/services/v2/ReportService"
                headers = {"Content-Type": "text/xml;charset=UTF-8", "SOAPAction": "runReport"}
                
                logger.info("Making SOAP request with working envelope...")
                logger.info(f"Service URL: {service_url}")
                logger.info(f"Report Path: {report_path}")
                logger.info(f"SOAP Envelope: {envelope[:500]}...")
                
                resp = requests.post(service_url, data=envelope.encode('utf-8'), headers=headers, timeout=300, verify=False)

                if resp.status_code == 200:
                    logger.info("SOAP request successful!")
                    
                    # Extract <reportBytes>
                    ns = {'soapenv': 'http://schemas.xmlsoap.org/soap/envelope/', 'v2': 'http://xmlns.oracle.com/oxp/service/v2'}
                    try:
                        root = ET.fromstring(resp.text)
                        rb = root.find('.//v2:reportBytes', ns)
                        if rb is None or not rb.text:
                            logger.error("No reportBytes in SOAP response")
                            return None, 0
                        csv_bytes = base64.b64decode(rb.text)
                    except Exception as parse_err:
                        logger.error(f"Failed parsing SOAP response: {parse_err}")
                        return None, 0

                    # Convert CSV bytes to Excel bytes
                    try:
                        import pandas as pd
                        import io
                        df = pd.read_csv(io.BytesIO(csv_bytes), delimiter='|', encoding='utf-8', dtype=str)
                        df = df.dropna(axis=1, how='all').fillna('')
                        out = io.BytesIO()
                        with pd.ExcelWriter(out, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name='Query Results')
                        excel_bytes = out.getvalue()
                        row_count = len(df)
                        logger.info(f"SOAP Excel generation successful: {row_count} rows")
                        return excel_bytes, row_count
                    except Exception as conv_err:
                        logger.error(f"Failed converting CSV to Excel: {conv_err}")
                        return None, 0
                else:
                    logger.error(f"SOAP HTTP error: {resp.status_code}")
                    logger.error(f"SOAP Response headers: {dict(resp.headers)}")
                    logger.error(f"SOAP Response content: {resp.text[:1000]}...")  # First 1000 chars
                    return None, 0
            except Exception as e:
                logger.error(f"SOAP raw POST flow failed: {e}")
                logger.error(f"Exception type: {type(e).__name__}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")

            # Fallback to zeep client if legacy is unavailable or failed
            session = Session()
            session.verify = False  # For self-signed certificates
            transport = Transport(session=session)
            client = Client(self.bip_wsdl_url, transport=transport)
            # Note: Some BIP WSDLs expose runReport instead of generateReport
            try:
                response = client.service.generateReport(
                    username=self.bip_username,
                    password=self.bip_password,
                    reportPath=report_path,  # Use decoded path
                    sqlQuery=safe_sql,
                    outputFormat='EXCEL'
                )
            except Exception:
                # Try runReport if generateReport is not present
                try:
                    import base64
                    b64_sql = base64.b64encode(safe_sql.encode('utf-8')).decode('ascii')
                    response = client.service.runReport(
                        reportRequest={
                            'reportAbsolutePath': report_path,  # Use decoded path
                            'attributeFormat': 'xlsx',
                            'sizeOfDataChunkDownload': -1,
                            'parameterNameValues': {
                                'item': [
                                    {
                                        'name': 'P_B64_CONTENT',
                                        'values': {
                                            'item': [b64_sql]
                                        }
                                    }
                                ]
                            }
                        },
                        userID=self.bip_username,
                        password=self.bip_password
                    )
                except Exception as e:
                    logger.error(f"SOAP Excel generation failed: {e}")
                    return None, 0

            if response is None:
                logger.warning("SOAP response empty or invalid")
                return None, 0

            # Normalize bytes from different response shapes
            try:
                report_bytes = None
                if hasattr(response, 'reportBytes'):
                    report_bytes = response.reportBytes
                elif hasattr(response, 'reportData'):
                    report_bytes = response.reportData
                excel_data = report_bytes.encode('utf-8') if isinstance(report_bytes, str) else report_bytes
            except Exception:
                logger.warning("Could not extract report bytes from SOAP response")
                return None, 0

            # Count rows (approximate)
            try:
                import io
                df = pd.read_excel(io.BytesIO(excel_data))
                row_count = len(df)
            except Exception:
                row_count = -1
            
            logger.info(f"SOAP Excel generation successful: {row_count} rows")
            return excel_data, row_count
                
        except Exception as e:
            logger.error(f"SOAP Excel generation failed: {e}")
            return None, 0
    
    def generate_excel_via_selenium(self, sql_query: str) -> tuple[bytes, int]:
        """Generate Excel report via Selenium automation."""
        try:
            logger.info("Attempting Selenium-based Excel generation...")
            
            # Remove semicolon at the end as BI Publisher doesn't accept it
            safe_sql = sql_query.rstrip(';').strip()
            if safe_sql != sql_query:
                logger.info("Removed semicolon from SQL query for BI Publisher compatibility")
            
            # Import BIP automator
            try:
                from bip_automator import BIPAutomator
            except ImportError as e:
                logger.error(f"BIP automator not available: {e}")
                return None, 0
            
            automator = BIPAutomator()
            excel_path = automator.automate_excel_generation(safe_sql)
            
            if excel_path and os.path.exists(excel_path):
                with open(excel_path, "rb") as f:
                    excel_bytes = f.read()
                
                # Count rows
                try:
                    df = pd.read_excel(excel_path)
                    row_count = len(df)
                except Exception as e:
                    logger.warning(f"Could not count rows: {e}")
                    row_count = -1
                
                logger.info(f"Selenium Excel generation successful: {row_count} rows")
                return excel_bytes, row_count
            else:
                logger.error(f"Excel file not found at: {excel_path}")
                return None, 0
                
        except Exception as e:
            logger.error(f"Selenium Excel generation failed: {e}")
            return None, 0
    
    def generate_simple_excel_fallback(self, sql_query: str) -> tuple[bytes, int]:
        """Generate simple Excel file with SQL query when all other methods fail."""
        try:
            logger.info("Generating simple Excel fallback with SQL query...")
            
            # Remove semicolon at the end as BI Publisher doesn't accept it
            safe_sql = sql_query.rstrip(';').strip()
            if safe_sql != sql_query:
                logger.info("Removed semicolon from SQL query for BI Publisher compatibility")
            
            # Create Excel workbook
            wb = Workbook()
            
            # SQL Query sheet
            query_sheet = wb.active
            query_sheet.title = "SQL Query"
            query_sheet['A1'] = "Generated SQL Query"
            query_sheet['A2'] = safe_sql
            query_sheet['A1'].font = Font(bold=True)
            query_sheet['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            query_sheet.column_dimensions['A'].width = 100
            
            # Instructions sheet
            instructions_sheet = wb.create_sheet("Instructions")
            instructions_sheet['A1'] = "Instructions for Manual Execution"
            instructions_sheet['A2'] = "1. Copy the SQL query from the 'SQL Query' sheet"
            instructions_sheet['A3'] = "2. Execute it in your Oracle database"
            instructions_sheet['A4'] = "3. Export results to Excel manually"
            instructions_sheet['A1'].font = Font(bold=True, color="FF0000")
            instructions_sheet.column_dimensions['A'].width = 80
            
            # Save to bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info("Simple Excel fallback generation successful")
            return output.getvalue(), 0  # 0 rows since no data
            
        except Exception as e:
            logger.error(f"Error generating simple Excel fallback: {e}")
            return None, 0
    
    def handle_sql_to_excel(self, sql_query: str) -> tuple[bytes, int]:
        """Main Excel generation handler with multiple fallback approaches."""
        start_time = time.time()
        logger.info(f"Excel generation started at: {datetime.now().strftime('%H:%M:%S')}")
        logger.info(f"SQL Query: {sql_query}")
        
        if not sql_query or "SELECT" not in sql_query.upper():
            logger.warning("No valid SQL query provided")
            st.error("No valid SQL query provided.")
            return None, 0
        
        try:
            with st.spinner("🔧 Generating Excel report..."):
                # Try SOAP-based approach first
                logger.info("Attempting SOAP-based Excel generation...")
                excel_data, row_count = self.generate_excel_via_soap(sql_query)
                
                if excel_data:
                    end_time = time.time()
                    duration = end_time - start_time
                    logger.info(f"SOAP Excel generation successful: {row_count} rows in {duration:.2f}s")
                    
                    if row_count == 0:
                        st.warning(f"⚠️ Query returned no data. An empty Excel file has been generated. (Time: {duration:.2f}s)")
                    else:
                        st.success(f"✅ Excel generation completed successfully! Generated {row_count} rows of data. (Time: {duration:.2f}s)")
                    
                    return excel_data, row_count
                
                # SOAP failed, try Selenium-based approach
                logger.info("SOAP failed, trying Selenium approach...")
                excel_data, row_count = self.generate_excel_via_selenium(sql_query)
                
                if excel_data:
                    end_time = time.time()
                    duration = end_time - start_time
                    logger.info(f"Selenium Excel generation successful: {row_count} rows in {duration:.2f}s")
                    
                    if row_count == 0:
                        st.warning(f"⚠️ Query returned no data. An empty Excel file has been generated. (Time: {duration:.2f}s)")
                    else:
                        st.success(f"✅ Excel generation completed successfully! Generated {row_count} rows of data. (Time: {duration:.2f}s)")
                    
                    return excel_data, row_count
                
                # All approaches failed, try simple fallback
                logger.info("SOAP and Selenium failed, trying simple fallback...")
                excel_data, row_count = self.generate_simple_excel_fallback(sql_query)
                
                if excel_data:
                    end_time = time.time()
                    duration = end_time - start_time
                    st.warning(f"⚠️ Using fallback Excel generation. Copy the SQL query and execute manually. (Time: {duration:.2f}s)")
                    logger.info(f"Simple fallback Excel generation successful in {duration:.2f}s")
                    return excel_data, row_count
                
                # All approaches failed
                end_time = time.time()
                duration = end_time - start_time
                st.error(f"Excel generation failed. (Time: {duration:.2f}s)")
                logger.error("All Excel generation approaches failed")
                return None, 0
                
        except Exception as e:
            end_time = time.time()
            duration = end_time - start_time
            st.error(f"Excel generation failed: {str(e)} (Time: {duration:.2f}s)")
            logger.error(f"Excel generation failed: {str(e)}")
            import traceback
            traceback.print_exc()
            return None, 0

# Global Excel generator instance
excel_generator = ExcelGenerator()